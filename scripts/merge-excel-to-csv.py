#!/usr/bin/env python3
"""Consolida todas as abas do Excel em um único CSV, ignorando MT e GO."""

import csv
import openpyxl

EXCEL_FILE = "Spiders funcionando .xlsx"
OUTPUT_CSV = "spiders_funcionando_consolidado.csv"
SHEETS_TO_IGNORE = {"MT", "GO"}

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    
    all_rows = []
    header = None
    
    for sheet_name in wb.sheetnames:
        if sheet_name in SHEETS_TO_IGNORE:
            continue
            
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            continue
            
        for i, row in enumerate(rows):
            # Take first 5 columns (id, Cidade, Status, Observações, Falhas)
            values = [v for v in row[:5]]
            # Pad if shorter
            while len(values) < 5:
                values.append(None)
            
            if i == 0:
                # Header row - add UF column
                if header is None:
                    header = ["UF"] + [str(v) if v is not None else "" for v in values]
                    all_rows.append(header)
            else:
                # Data row - add sheet name as UF
                str_values = [str(v) if v is not None else "" for v in values]
                all_rows.append([sheet_name] + str_values)
    
    included_sheets = [s for s in wb.sheetnames if s not in SHEETS_TO_IGNORE]
    wb.close()
    
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(all_rows)
    
    print(f"Arquivo salvo: {OUTPUT_CSV}")
    print(f"Total de linhas (incluindo cabeçalho): {len(all_rows)}")
    print(f"Abas incluídas: {included_sheets}")

if __name__ == "__main__":
    main()
